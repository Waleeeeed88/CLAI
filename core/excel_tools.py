"""Excel test plan generator — creates formatted .xlsx test plans.

Provides a ``create_test_plan_excel`` tool that agents (especially QA)
can call to produce structured test plan spreadsheets in the workspace.

Requires ``openpyxl``.
"""
from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from core.tool_registry import ToolParameter, ToolRegistry

logger = logging.getLogger(__name__)


def _create_test_plan(
    file_path: str,
    suite_name: str,
    test_cases: str,
    workspace_root: Optional[Path] = None,
) -> str:
    """Create a formatted Excel test plan.

    Args:
        file_path: Relative path for the .xlsx file in the workspace.
        suite_name: Name/title of the test suite.
        test_cases: JSON string — list of objects with keys:
            id, title, steps, expected_result, priority, status, category
        workspace_root: Workspace root path (resolved externally).

    Returns:
        Success or error message.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "Error: openpyxl is not installed. Run: pip install openpyxl"

    # Parse test cases
    try:
        cases = json.loads(test_cases) if isinstance(test_cases, str) else test_cases
    except json.JSONDecodeError as e:
        return f"Error parsing test_cases JSON: {e}"

    if not isinstance(cases, list):
        return "Error: test_cases must be a JSON array of objects."

    # Resolve output path
    if workspace_root is None:
        from config import get_settings
        workspace_root = get_settings().workspace_path

    clean_path = file_path.lstrip("/").lstrip("\\")
    if not clean_path.endswith(".xlsx"):
        clean_path += ".xlsx"
    full_path = (workspace_root / clean_path).resolve()
    try:
        full_path.relative_to(workspace_root)
    except ValueError:
        return f"Error: Path '{file_path}' escapes workspace sandbox."

    full_path.parent.mkdir(parents=True, exist_ok=True)

    # ── Build workbook ───────────────────────────────────────────────

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Plan"

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    priority_fills = {
        "high": PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
        "critical": PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
        "medium": PatternFill(start_color="FFFDE0", end_color="FFFDE0", fill_type="solid"),
        "low": PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid"),
    }

    status_fills = {
        "pass": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "passed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "fail": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "failed": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "blocked": PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid"),
        "not run": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        "not_run": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    }

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Test Plan: {suite_name}"
    title_cell.font = Font(bold=True, size=16, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Header row
    headers = ["ID", "Title", "Steps", "Expected Result", "Priority", "Status", "Category"]
    col_widths = [8, 35, 50, 40, 12, 12, 18]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    for row_idx, case in enumerate(cases, start=4):
        values = [
            case.get("id", row_idx - 3),
            case.get("title", ""),
            case.get("steps", ""),
            case.get("expected_result", ""),
            case.get("priority", "medium"),
            case.get("status", "not run"),
            case.get("category", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Apply priority fill
            if col_idx == 5:
                fill = priority_fills.get(str(val).lower())
                if fill:
                    cell.fill = fill

            # Apply status fill
            if col_idx == 6:
                fill = status_fills.get(str(val).lower())
                if fill:
                    cell.fill = fill

    # Freeze panes
    ws.freeze_panes = "A4"

    # Auto-filter
    ws.auto_filter.ref = f"A3:G{3 + len(cases)}"

    # Save
    wb.save(str(full_path))
    return f"Test plan created: {clean_path} ({len(cases)} test cases)"


def build_excel_registry(workspace_root: Optional[Path] = None) -> ToolRegistry:
    """Build a ToolRegistry with the Excel test plan tool."""
    registry = ToolRegistry()

    def handler(file_path: str, suite_name: str, test_cases: str) -> str:
        return _create_test_plan(file_path, suite_name, test_cases, workspace_root)

    registry.register(
        name="create_test_plan_excel",
        description=(
            "Create a formatted Excel (.xlsx) test plan document in the workspace. "
            "Provide test cases as a JSON array where each object has keys: "
            "id, title, steps, expected_result, priority (high/medium/low), "
            "status (not run/pass/fail/blocked), category."
        ),
        parameters=[
            ToolParameter(
                "file_path", "string",
                "Relative path for the .xlsx file (e.g. 'myproject/test_plan.xlsx')",
                required=True,
            ),
            ToolParameter(
                "suite_name", "string",
                "Name/title of the test suite",
                required=True,
            ),
            ToolParameter(
                "test_cases", "string",
                'JSON array of test case objects, e.g. [{"id":"TC001","title":"Login test","steps":"1. Open app\\n2. Enter credentials","expected_result":"User logged in","priority":"high","status":"not run","category":"auth"}]',
                required=True,
            ),
        ],
        handler=handler,
    )

    return registry
